# Importar librerias necesarias
import cv2
import os
from openpyxl import load_workbook
import unicodedata
import re
import time

# Iniciar el contador de tiempo
start_time = time.time()

# Conectar a la base de datos de nombres (Excel)
wb = load_workbook('base_datos_nombres/Base_datos_nombres.xlsx')
sheet = wb.active

# Carpeta de fotos originales
carpeta_originales = 'Base_datos_fotos/'

# Carpeta de fotos recortadas
carpeta_recortes = 'Base_datos_fotos_recortadas/'

# Lista para almacenar los nombres que no hicieron match o no se detectaron rostros
nombres_errores = []

# Contador para llevar el registro de elementos recortados y guardados
elementos_recortados = 0

# Obtener la lista de fotos en la carpeta donde se recortan
fotos_originales = os.listdir(carpeta_originales)

# Expresión regular para encontrar el primer grupo de dígitos en el nombre del archivo
pattern = re.compile(r'\d+')

for foto_original in fotos_originales:
    # Obtener el *Person ID de la foto original
    match = pattern.search(foto_original)
    if match:
        person_id_foto = int(match.group())
    else:
        # Si no se encuentra un número, agregar el nombre a la lista de errores y continuar con la siguiente foto
        nombres_errores.append((foto_original, "Error: No se reconoció el rostro"))
        continue

    # Buscar el nombre correspondiente en la hoja de Excel
    nombre_asociado = None
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Validar si el Person ID de la foto está dentro del rango de Person IDs en la hoja de Excel
        if person_id_foto in range(row[0], row[1] + 1) if row[1] is not None else (row[0] == person_id_foto):
            nombre_asociado = row[2]  # *Person Name
            break

    if nombre_asociado:
        # Normalizar los caracteres para evitar problemas con acentos y caracteres especiales
        nuevo_nombre = unicodedata.normalize('NFKD', f"{person_id_foto}_{nombre_asociado}.jpg").encode('ascii', 'ignore').decode('utf-8')

        # Cargar la imagen original
        image_path = os.path.join(carpeta_originales, foto_original)
        image = cv2.imread(image_path)

        # Convertir la imagen a escala de grises
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        # Cargar el clasificador de Haar para la detección de caras
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

        # Detectar los rostros en la imagen
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)

        if len(faces) > 0:  # Verificar si se detectaron rostros
            # Crear el directorio si no existe
            os.makedirs(carpeta_recortes, exist_ok=True)

            # Resto del código para recortar y guardar los rostros
            for i, (x, y, w, h) in enumerate(faces):
                # Expandir la región superior e inferior
                top_expansion_pixels = int(0.15 * image.shape[0])  # 15% de la altura de la imagen
                bottom_expansion_pixels = int(0.1 * image.shape[0])  # 10% de la altura de la imagen
                y = max(0, y - top_expansion_pixels)
                h = min(image.shape[0], h + top_expansion_pixels + bottom_expansion_pixels)

                # Resto del código para recortar y guardar los rostros
                face = image[y:y+h, x:x+w]
                nuevo_nombre_recortada = f"{nuevo_nombre.split('.')[0]}_{i}.jpg"
                cv2.imwrite(os.path.join(carpeta_recortes, nuevo_nombre_recortada), face)
                elementos_recortados += 1
        else:
            # Agregar el nombre original a la lista de nombres sin match o con error
            nombres_errores.append((foto_original, "Error: No se detectó rostro"))
    else:
        # Agregar el nombre original a la lista de nombres sin match o con error
        nombres_errores.append((foto_original, "Error: No se encontró coincidencia con *Person ID"))

# Cerrar el archivo Excel
wb.close()

# Guardar los nombres sin match o con error en un archivo de texto
with open('nombres_errores.txt', 'w') as file:
    for nombre, error in nombres_errores:
        file.write(f"{nombre}: {error}\n")

# Imprimir la cantidad de elementos recortados, guardados. Calcular y mostrar el tiempo transcurrido
elapsed_time = time.time() - start_time
print(f"Proceso completado. Elementos recortados y guardados: {elementos_recortados}. Tiempo transcurrido: {elapsed_time:.2f} segundos")












# def crop_faces(image_path, top_expansion_factor=0.15, bottom_expansion_factor=0.1):
#     # Cargar la imagen
#     image = cv2.imread(image_path)

#     # Convertir la imagen a escala de grises
#     gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

#     # Cargar el clasificador de Haar para la detección de caras
#     face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

#     # Detectar los rostros en la imagen
#     faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)

#     # Expandir la región superior e inferior
#     top_expansion_pixels = int(top_expansion_factor * image.shape[0])  # 15% de la altura de la imagen
#     bottom_expansion_pixels = int(bottom_expansion_factor * image.shape[0])  # 10% de la altura de la imagen

#     # Recortar y guardar cada rostro detectado
#     for i, (x, y, w, h) in enumerate(faces):
#         # Expandir la región superior e inferior
#         y = max(0, y - top_expansion_pixels)
#         h = min(image.shape[0], h + top_expansion_pixels + bottom_expansion_pixels)

#         face = image[y:y+h, x:x+w]
#         cv2.imwrite(f"Base_datos_fotos_recortadas/face_{i+1}_{os.path.basename(image_path)}", face)

# # Directorio base
# base_dir = "/Users/homet/proyecto1/baseDatosUsi/foto_recortadas/Base_datos_fotos/"

# # Iterar sobre las imágenes en el directorio
# for image_file in os.listdir(base_dir):
#     if image_file.endswith(('.jpg', '.jpeg', '.png')):
#         image_path = os.path.join(base_dir, image_file)
#         crop_faces(image_path)

# print("Proceso completado.")



